"""
uploader.py — Google Sheets upload + Excel fallback (single sheet)

SAFETY FEATURES:
  1. Upsert mode — updates changed cells in place, appends new rows. NEVER clears/deletes.
  2. Row count validation — reads row count BEFORE and AFTER, aborts if mismatch.
  3. Local backup — saves a snapshot of the existing sheet BEFORE any write.
  4. get_all_values — uses raw cell reads instead of get_all_records (more reliable).
  5. Post-write verification — confirms total rows after write matches expected count.
  6. Targeted cell writes — only touches cells whose values actually changed.
"""

import os
from datetime import date, datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import (
    CREDENTIALS_FILE, SHEET_AD_PERFORMANCE,
    OUTPUT_DIR, DEDUP_KEYS_AD
)

TODAY = date.today().strftime("%m/%d/%Y")
BACKUP_DIR = os.path.join(OUTPUT_DIR, "backups")


# ── SAFE READ ─────────────────────────────────────────────────────────────────

def _safe_read_sheet(ws):
    """
    Read all data from a worksheet using get_all_values() (raw cells)
    instead of get_all_records() which silently truncates on large sheets.
    Returns a DataFrame.
    """
    all_values = ws.get_all_values()
    if not all_values or len(all_values) < 2:
        return pd.DataFrame()

    headers = all_values[0]
    data = all_values[1:]
    df = pd.DataFrame(data, columns=headers)

    # Drop fully empty rows (all blank strings)
    df = df[df.apply(lambda r: any(str(v).strip() != '' for v in r), axis=1)]
    return df


# ── LOCAL BACKUP ──────────────────────────────────────────────────────────────

def _backup_existing(existing_df, tab_name):
    """
    Save a local Excel backup of the existing sheet data BEFORE any changes.
    Returns the backup filepath.
    """
    os.makedirs(BACKUP_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"{tab_name}_backup_{timestamp}.xlsx")
    existing_df.to_excel(backup_path, index=False)
    print(f"         💾 Backup saved → {backup_path} ({len(existing_df):,} rows)")
    return backup_path


# ── UPLOAD ────────────────────────────────────────────────────────────────────

def upload_to_sheets(combined, sheet_id):
    if not sheet_id or sheet_id == "YOUR_GOOGLE_SHEET_ID_HERE":
        print("      ⚠️  No Google Sheet ID configured — skipping upload.")
        return False
    if not os.path.exists(CREDENTIALS_FILE):
        print(f"      ⚠️  credentials.json not found — skipping upload.")
        return False
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds  = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
        client = gspread.authorize(creds)
        sheet  = client.open_by_key(sheet_id)
        _write_tab(sheet, SHEET_AD_PERFORMANCE, combined, DEDUP_KEYS_AD)
        print(f"      ✅ Google Sheet updated.")
        print(f"         https://docs.google.com/spreadsheets/d/{sheet_id}")
        return True
    except ImportError:
        print("      ❌ gspread not installed. Run: pip install gspread google-auth")
        return False
    except Exception as e:
        import traceback
        print(f"      ❌ Google Sheets upload failed: {e}")
        traceback.print_exc()
        return False


def _write_tab(sheet, tab_name, new_df, dedup_keys):
    """
    UPSERT write: updates changed rows IN PLACE, appends truly new rows.
    Never clears the sheet. Never reduces row count.

    1. Read existing data safely
    2. Back up locally
    3. Validate row count
    4. Split incoming data into UPDATES (key exists) vs APPENDS (key is new)
    5. Apply updates via targeted cell writes (only changed cells)
    6. Append new rows below existing data
    7. Post-write verification
    """

    # ── Columns that can be updated when a matching row exists ─────────
    UPDATABLE_COLS = ["Impressions", "Clicks", "CTR", "Spend (AUD)", "QL", "FT"]

    # ── Step 1: Read existing data ────────────────────────────────────────
    is_new_sheet = False
    try:
        ws = sheet.worksheet(tab_name)
        existing_df = _safe_read_sheet(ws)
        existing_row_count = ws.row_count
        print(f"         📊 Existing sheet: {len(existing_df):,} data rows ({existing_row_count:,} total rows incl. header/blanks)")
    except Exception:
        ws = sheet.add_worksheet(title=tab_name, rows=50000, cols=20)
        existing_df = pd.DataFrame()
        is_new_sheet = True
        print(f"         📊 New sheet created: {tab_name}")

    # ── Step 2: Backup existing data locally ──────────────────────────────
    if not existing_df.empty:
        _backup_existing(existing_df, tab_name)

    # ── Step 3: Validate read integrity ───────────────────────────────────
    if not existing_df.empty:
        col_a = ws.col_values(1)
        non_empty_a = len([v for v in col_a[1:] if str(v).strip() != ''])
        if abs(non_empty_a - len(existing_df)) > 5:
            print(f"         ❌ SAFETY ABORT: Row count mismatch!")
            print(f"            get_all_values returned {len(existing_df):,} rows")
            print(f"            Column A has {non_empty_a:,} non-empty cells")
            print(f"            Difference too large — refusing to write to avoid data loss.")
            print(f"            Backup was saved. Please check the sheet manually.")
            return
        print(f"         ✅ Read integrity verified ({len(existing_df):,} rows)")

    # ── Step 4: Prepare new data & split into updates vs appends ──────────
    new_df = new_df.copy()
    new_df["Date"] = pd.to_datetime(new_df["Date"]).dt.strftime("%m/%d/%Y")

    if not existing_df.empty:
        # Normalise existing dates to match strftime format (with leading zeros)
        # Sheets may return "1/5/2026" but strftime produces "01/05/2026"
        existing_df["Date"] = pd.to_datetime(existing_df["Date"], format="mixed", dayfirst=False).dt.strftime("%m/%d/%Y")

        # ── Build row-type-aware composite key ────────────────────────────
        # QL/FT rows (no impressions/clicks/spend) → match on Date+Country+Channel+Channel_Group
        # Ad rows (have impressions/clicks/spend)  → match on Date+Country+Channel+Campaign+Creative

        AD_KEYS   = ["Date", "Country", "Channel", "Campaign", "Creative"]
        QLFT_KEYS = ["Date", "Country", "Channel", "Channel_Group"]

        def _normalise_val(v):
            """Normalise a value for key matching: None/NaN/empty → ''"""
            s = str(v).strip()
            if s in ('', 'nan', 'None', 'NaN', 'none', 'null'):
                return ''
            return s

        def _is_qlft_row(row):
            """A row is QL/FT if it has no Impressions AND no Spend."""
            imp = _normalise_val(row.get("Impressions", ""))
            spn = _normalise_val(row.get("Spend (AUD)", ""))
            return imp == '' and spn == ''

        def make_key(row):
            keys = QLFT_KEYS if _is_qlft_row(row) else AD_KEYS
            return "||".join(_normalise_val(row.get(k, "")) for k in keys)

        existing_df["_key"] = existing_df.apply(make_key, axis=1)
        new_df["_key"]      = new_df.apply(make_key, axis=1)

        # Map existing key → sheet row number (1-indexed, +1 for header)
        key_to_sheet_row = {}
        for idx, key in existing_df["_key"].items():
            key_to_sheet_row[key] = idx + 2  # +1 for 0-index, +1 for header row

        # Map column name → sheet column index (1-indexed)
        headers = existing_df.columns.tolist()
        col_name_to_idx = {name: i + 1 for i, name in enumerate(headers) if name != "_key"}

        # Split: rows whose key already exists → updates; rest → appends
        existing_key_set = set(existing_df["_key"])
        updates_mask     = new_df["_key"].isin(existing_key_set)
        updates_df       = new_df[updates_mask].copy()
        appends_df       = new_df[~updates_mask].copy()

        # ── Apply updates: targeted cell writes only for changed values ───
        n_updated = 0
        cell_updates = []  # Collect all (row, col, value) for batch write

        for _, new_row in updates_df.iterrows():
            key = new_row["_key"]
            if key not in key_to_sheet_row:
                continue
            sheet_row = key_to_sheet_row[key]
            ex_idx = sheet_row - 2  # back to df index

            row_changed = False
            for col in UPDATABLE_COLS:
                if col not in col_name_to_idx:
                    continue
                new_val = new_row.get(col)
                old_val = existing_df.at[ex_idx, col] if col in existing_df.columns else ""

                # Skip if new value is empty/NaN
                if pd.isna(new_val) or str(new_val).strip() in ("", "nan", "None"):
                    continue

                # Check if actually different
                old_str = str(old_val).strip()
                changed = False
                if old_str in ("", "nan", "None", "0"):
                    changed = True
                else:
                    try:
                        if abs(float(new_val) - float(old_str)) > 0.0001:
                            changed = True
                    except (ValueError, TypeError):
                        if str(new_val).strip() != old_str:
                            changed = True

                if changed:
                    sheet_col = col_name_to_idx[col]
                    # Format the value for the sheet
                    try:
                        cell_val = float(new_val) if pd.notna(new_val) else ""
                    except (ValueError, TypeError):
                        cell_val = str(new_val) if pd.notna(new_val) else ""
                    cell_updates.append((sheet_row, sheet_col, cell_val))
                    row_changed = True

            if row_changed:
                n_updated += 1

        # Write all cell updates in batches
        if cell_updates:
            # gspread batch_update format: list of {'range': 'A1', 'values': [[val]]}
            batch = []
            for row, col, val in cell_updates:
                cell_ref = f"{get_column_letter(col)}{row}"
                batch.append({'range': cell_ref, 'values': [[val]]})

            # gspread limit: ~60k cells per batch_update call
            CELL_BATCH = 500
            for i in range(0, len(batch), CELL_BATCH):
                ws.batch_update(batch[i:i + CELL_BATCH])

            print(f"         ✏️  Updated {n_updated:,} existing rows ({len(cell_updates):,} cells changed)")
        else:
            print(f"         ✓ No existing rows needed updating")

        # ── Prepare appends ───────────────────────────────────────────────
        appends_df = appends_df.drop(columns=["_key"], errors="ignore")
        appends_df["Date_Added"] = TODAY

        if len(appends_df) == 0 and n_updated == 0:
            print(f"         ✓ {tab_name}: no changes — nothing added or updated (existing: {len(existing_df):,} rows)")
            return
        elif len(appends_df) == 0:
            print(f"         ✓ {tab_name}: no new rows to append")
        else:
            print(f"         ✓ {tab_name}: {len(appends_df):,} new rows to append")

        new_rows = appends_df

    else:
        # Fresh sheet — all rows are new
        new_rows = new_df.drop(columns=["_key"], errors="ignore").copy()
        new_rows["Date_Added"] = TODAY
        n_updated = 0
        print(f"         ✓ {tab_name}: {len(new_rows):,} rows to write (fresh sheet)")

    # ── Step 5: Write appends ─────────────────────────────────────────────

    if len(new_rows) > 0:
        if is_new_sheet or existing_df.empty:
            # Fresh sheet: write header + all rows
            all_rows = [new_rows.columns.tolist()] + new_rows.fillna("").values.tolist()
            required_rows = len(all_rows) + 500
            ws.resize(rows=required_rows, cols=len(new_rows.columns))

            BATCH_SIZE = 400
            for i in range(0, len(all_rows), BATCH_SIZE):
                chunk = all_rows[i:i + BATCH_SIZE]
                start_row = i + 1
                ws.update(chunk, f"A{start_row}")

            print(f"         ✅ Wrote {len(new_rows):,} rows to fresh sheet")

        else:
            # APPEND below existing data
            append_start = len(existing_df) + 2  # +1 header, +1 next empty row
            rows_needed = append_start + len(new_rows) + 100
            if rows_needed > ws.row_count:
                ws.resize(rows=rows_needed, cols=len(new_rows.columns))

            append_data = new_rows.fillna("").values.tolist()

            BATCH_SIZE = 400
            for i in range(0, len(append_data), BATCH_SIZE):
                chunk = append_data[i:i + BATCH_SIZE]
                row_pos = append_start + i
                ws.update(chunk, f"A{row_pos}")

            print(f"         ✅ Appended {len(new_rows):,} rows starting at row {append_start}")

    # ── Step 6: Post-write verification ───────────────────────────────────
    import time
    time.sleep(2)
    col_a_after = ws.col_values(1)
    final_count = len([v for v in col_a_after[1:] if str(v).strip() != ''])
    n_appended = len(new_rows) if len(new_rows) > 0 else 0
    expected = (len(existing_df) + n_appended) if not existing_df.empty else n_appended

    print(f"         📊 Final: {final_count:,} rows ({n_updated:,} updated, {n_appended:,} appended)")

    if abs(final_count - expected) <= 5:
        print(f"         ✅ Post-write verification passed")
    else:
        print(f"         ⚠️  Post-write count mismatch: expected ~{expected:,}, found {final_count:,}")
        print(f"            Backup was saved — check manually if needed.")


# ── POST-RUN SNAPSHOT ─────────────────────────────────────────────────────────

def save_run_snapshot(sheet_id):
    """
    Download the FULL Google Sheet and save as a dated Excel file.
    Called after every successful upload for version history.
    Files: ./output/snapshots/Ad_Performance_YYYYMMDD_HHMMSS.xlsx
    """
    SNAPSHOT_DIR = os.path.join(OUTPUT_DIR, "snapshots")
    os.makedirs(SNAPSHOT_DIR, exist_ok=True)

    try:
        import gspread
        from google.oauth2.service_account import Credentials
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds  = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
        client = gspread.authorize(creds)
        sheet  = client.open_by_key(sheet_id)
        ws     = sheet.worksheet(SHEET_AD_PERFORMANCE)

        df = _safe_read_sheet(ws)
        if df.empty:
            print("      ⚠️  Snapshot skipped — sheet is empty.")
            return None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath  = os.path.join(SNAPSHOT_DIR, f"Ad_Performance_{timestamp}.xlsx")

        # Convert numeric columns before saving
        for col in ["Impressions", "Clicks", "Spend (AUD)", "CTR", "QL", "FT"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        df.to_excel(filepath, index=False)
        print(f"      💾 Snapshot saved → {filepath} ({len(df):,} rows)")
        return filepath

    except Exception as e:
        print(f"      ⚠️  Snapshot failed: {e}")
        return None


# ── EXCEL FALLBACK ────────────────────────────────────────────────────────────

def fallback_to_excel(combined, filepath):
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    UPDATABLE_COLS = ["Impressions", "Clicks", "CTR", "Spend (AUD)", "QL", "FT"]

    out = combined.copy()
    out["Date"] = pd.to_datetime(out["Date"]).dt.strftime("%m/%d/%Y")

    if os.path.exists(filepath):
        try:
            existing = pd.read_excel(filepath, sheet_name=SHEET_AD_PERFORMANCE, dtype=str)

            # Normalise dates in existing file to match strftime format
            existing["Date"] = pd.to_datetime(existing["Date"], format="mixed", dayfirst=False).dt.strftime("%m/%d/%Y")

            # ── Build row-type-aware composite key ──────────────────────
            AD_KEYS   = ["Date", "Country", "Channel", "Campaign", "Creative"]
            QLFT_KEYS = ["Date", "Country", "Channel", "Channel_Group"]

            def _normalise_val(v):
                s = str(v).strip()
                if s in ('', 'nan', 'None', 'NaN', 'none', 'null'):
                    return ''
                return s

            def _is_qlft_row(row):
                imp = _normalise_val(row.get("Impressions", ""))
                spn = _normalise_val(row.get("Spend (AUD)", ""))
                return imp == '' and spn == ''

            def make_key(row):
                keys = QLFT_KEYS if _is_qlft_row(row) else AD_KEYS
                return "||".join(_normalise_val(row.get(k, "")) for k in keys)

            existing["_key"] = existing.apply(make_key, axis=1)
            out["_key"]      = out.apply(make_key, axis=1)

            existing_key_set = set(existing["_key"])
            updates_mask     = out["_key"].isin(existing_key_set)
            updates_df       = out[updates_mask].copy()
            appends_df       = out[~updates_mask].copy()

            # ── Apply updates to existing rows ────────────────────────────
            n_updated = 0
            if not updates_df.empty:
                key_to_idx = {}
                for idx, key in existing["_key"].items():
                    key_to_idx[key] = idx

                for _, new_row in updates_df.iterrows():
                    key = new_row["_key"]
                    if key not in key_to_idx:
                        continue
                    ex_idx = key_to_idx[key]
                    changed = False
                    for col in UPDATABLE_COLS:
                        if col not in out.columns or col not in existing.columns:
                            continue
                        new_val = new_row.get(col)
                        old_val = existing.at[ex_idx, col]
                        if pd.isna(new_val) or str(new_val).strip() in ("", "nan", "None"):
                            continue
                        old_str = str(old_val).strip()
                        if old_str in ("", "nan", "None", "0"):
                            changed = True
                        else:
                            try:
                                if abs(float(new_val) - float(old_str)) > 0.0001:
                                    changed = True
                            except (ValueError, TypeError):
                                if str(new_val).strip() != old_str:
                                    changed = True
                        if changed:
                            existing.at[ex_idx, col] = new_val
                    if changed:
                        n_updated += 1

            # ── Append truly new rows ─────────────────────────────────────
            appends_df = appends_df.drop(columns=["_key"], errors="ignore")
            appends_df["Date_Added"] = TODAY
            existing = existing.drop(columns=["_key"], errors="ignore")

            # SAFETY: merged must never have fewer rows than existing
            out = pd.concat([existing, appends_df], ignore_index=True).sort_values(DEDUP_KEYS_AD).reset_index(drop=True)
            print(f"      ↻  Upsert: {n_updated:,} rows updated, +{len(appends_df):,} new rows appended")

        except Exception as e:
            print(f"      ⚠️  Could not read existing file ({e}) — overwriting.")
            out = out.drop(columns=["_key"], errors="ignore")
            out["Date_Added"] = TODAY
    else:
        out["Date_Added"] = TODAY

    for col in ["Impressions","Clicks","Spend (AUD)","CTR","QL","FT"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")

    _write_excel(out, filepath)
    print(f"      ✅ Excel saved → {filepath}")
    print(f"         Total rows: {len(out):,}")


def _write_excel(df, filepath):
    wb = Workbook()
    HEADER_FILL = PatternFill("solid", start_color="1F3864")
    ALT_FILL    = PatternFill("solid", start_color="EEF2F7")
    WHITE_FILL  = PatternFill("solid", start_color="FFFFFF")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    DATA_FONT   = Font(name="Arial", size=10)
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left",   vertical="center")
    RIGHT       = Alignment(horizontal="right",  vertical="center")
    thin        = Side(style="thin", color="D0D0D0")
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.active
    ws.title = SHEET_AD_PERFORMANCE
    ws.freeze_panes = "A2"

    headers    = ["Date","Country","Channel","Campaign","Creative","Impressions","Clicks","CTR","Spend (AUD)","QL","FT","Channel_Group","Date_Added"]
    col_widths = [14, 10, 22, 48, 24, 14, 12, 10, 14, 10, 10, 16, 14]

    ws.row_dimensions[1].height = 22
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(1, ci, h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        ws.row_dimensions[ri].height = 18

        def c(ci, val, align=RIGHT):
            cell = ws.cell(ri, ci, val)
            cell.font = DATA_FONT; cell.fill = fill
            cell.alignment = align
            return cell

        c(1,  str(row.Date), LEFT)
        c(2,  str(row.Country)       if pd.notna(row.Country)       else "", LEFT)
        c(3,  str(row.Channel)       if pd.notna(row.Channel)       else "", LEFT)
        c(4,  str(row.Campaign)      if pd.notna(row.Campaign)      else "", LEFT)
        impr = getattr(row,"Impressions",None)
        c(5,  float(impr) if pd.notna(impr) else "")
        clks = getattr(row,"Clicks",None)
        c(6,  float(clks) if pd.notna(clks) else "")
        ctr  = getattr(row,"CTR",None)
        if pd.notna(ctr):
            cell = c(7, round(float(ctr)*100, 4))
            cell.number_format = '0.00"%"'
        else:
            c(7, "")
        spnd = getattr(row,"Spend (AUD)",None)
        c(8,  float(spnd) if pd.notna(spnd) else "")
        ql   = getattr(row,"QL",None)
        c(9,  int(float(ql)) if pd.notna(ql) and str(ql) != '' else "")
        ft   = getattr(row,"FT",None)
        c(10, int(float(ft)) if pd.notna(ft) and str(ft) != '' else "")
        c(11, str(row.Channel_Group) if pd.notna(row.Channel_Group) else "", LEFT)
        da   = getattr(row,"Date_Added",None)
        c(12, str(da) if pd.notna(da) and str(da) not in ('', 'nan', 'None') else "", LEFT)

    for ri in range(2, len(df)+2):
        ws.cell(ri,5).number_format = "#,##0"
        ws.cell(ri,6).number_format = "#,##0"
        ws.cell(ri,8).number_format = "#,##0.00"
        ws.cell(ri,9).number_format = "#,##0"
        ws.cell(ri,10).number_format = "#,##0"

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wb.save(filepath)